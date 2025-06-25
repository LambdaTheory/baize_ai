#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel导出模块
用于将AI图片信息导出为Excel表格，支持图片嵌入
"""

import os
import json
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        
    def export_records(self, records: List[Dict[str, Any]], output_path: str, 
                      include_images: bool = True, max_image_size: tuple = (150, 150)) -> bool:
        """
        导出记录到Excel文件
        
        Args:
            records: 记录列表
            output_path: 输出文件路径
            include_images: 是否包含图片
            max_image_size: 图片最大尺寸 (width, height)
            
        Returns:
            bool: 是否成功导出
        """
        temp_files_to_cleanup = []
        try:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "AI图片信息"
            
            # 设置表头
            self._setup_headers(include_images)
            
            # 填充数据，并收集临时文件列表
            temp_files_to_cleanup = self._fill_data(records, include_images, max_image_size)
            
            # 设置样式
            self._apply_styles(include_images)
            
            # 保存文件（这时openpyxl才真正读取图片文件）
            self.workbook.save(output_path)
            
            # 保存成功后清理临时文件
            self._cleanup_temp_files(temp_files_to_cleanup)
            
            return True
            
        except Exception as e:
            print(f"导出Excel失败: {e}")
            import traceback
            traceback.print_exc()
            # 即使出错也要清理临时文件
            self._cleanup_temp_files(temp_files_to_cleanup)
            return False
    
    def _setup_headers(self, include_images: bool):
        """设置表头"""
        headers = []
        
        if include_images:
            headers.append("图片")
            
        headers.extend([
            "文件名", "自定义名称", "正向提示词", "负向提示词", 
            "模型", "采样器", "采样步数", "CFG缩放", "种子", 
            "LoRA信息", "标签", "备注", "创建时间"
        ])
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _fill_data(self, records: List[Dict[str, Any]], include_images: bool, max_image_size: tuple) -> List[str]:
        """填充数据"""
        current_row = 2
        temp_files_to_cleanup = []  # 记录需要清理的临时文件
        
        for record in records:
            col = 1
            
            # 处理图片
            if include_images:
                image_path = record.get('file_path', '')
                # 先在图片列写入占位符
                img_cell = self.worksheet.cell(row=current_row, column=col, value="")
                
                if image_path and os.path.exists(image_path):
                    try:
                        # 调整图片大小并插入
                        resized_image_path = self._resize_image(image_path, max_image_size)
                        if resized_image_path and os.path.exists(resized_image_path):
                            # 使用更安全的方式处理图片
                            try:
                                img = Image(resized_image_path)
                                img.anchor = f"{get_column_letter(col)}{current_row}"
                                self.worksheet.add_image(img)
                                
                                # 设置行高以适应图片
                                self.worksheet.row_dimensions[current_row].height = max_image_size[1] * 0.8
                                
                                print(f"成功添加图片: {resized_image_path}")
                                
                                # 记录临时文件，稍后批量清理
                                if resized_image_path != image_path:
                                    temp_files_to_cleanup.append(resized_image_path)
                            except Exception as img_error:
                                print(f"openpyxl添加图片失败 {resized_image_path}: {img_error}")
                                # 清理失败的临时文件
                                if resized_image_path != image_path and os.path.exists(resized_image_path):
                                    try:
                                        os.remove(resized_image_path)
                                    except:
                                        pass
                                # 显示文件名
                                img_cell.value = os.path.basename(image_path)
                        else:
                            # 调整大小失败，显示文件名
                            img_cell.value = os.path.basename(image_path)
                    except Exception as e:
                        print(f"处理图片失败 {image_path}: {e}")
                        import traceback
                        traceback.print_exc()
                        # 图片处理失败时显示文件名
                        img_cell.value = os.path.basename(image_path)
                else:
                    # 图片不存在时显示提示
                    img_cell.value = "图片不存在"
                
                col += 1
            
            # 填充其他数据
            seed_value = record.get('seed', '')
            if seed_value and str(seed_value).isdigit():
                seed_value = str(seed_value)  # 强制转为字符串避免科学计数法
            
            data_values = [
                os.path.basename(record.get('file_path', '')),
                record.get('custom_name', ''),
                record.get('prompt', ''),
                record.get('negative_prompt', ''),
                record.get('model', ''),
                record.get('sampler', ''),
                record.get('steps', ''),
                record.get('cfg_scale', ''),
                seed_value,
                self._format_lora_info(record.get('lora_info', '')),
                record.get('tags', ''),
                record.get('notes', ''),
                record.get('created_at', '')[:19] if record.get('created_at') else ''
            ]
            
            for i, value in enumerate(data_values):
                cell = self.worksheet.cell(row=current_row, column=col, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # 种子列设置为文本格式（第9个数据列）
                if i == 8:  # 种子是data_values的第9个元素（索引8）
                    cell.number_format = '@'  # 文本格式
                    
                col += 1
            
            current_row += 1
        
        # 返回需要清理的临时文件列表
        return temp_files_to_cleanup
    
    def _resize_image(self, image_path: str, max_size: tuple) -> str:
        """调整图片大小"""
        try:
            # 检查原始文件是否存在
            if not os.path.exists(image_path):
                print(f"原始图片文件不存在: {image_path}")
                return None
                
            with PILImage.open(image_path) as img:
                # 计算缩放比例
                width_ratio = max_size[0] / img.width
                height_ratio = max_size[1] / img.height
                ratio = min(width_ratio, height_ratio)
                
                if ratio >= 1:
                    # 图片已经足够小，直接返回原路径
                    return image_path
                
                # 计算新尺寸
                new_width = int(img.width * ratio)
                new_height = int(img.height * ratio)
                
                # 调整大小
                resized_img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
                
                # 生成临时文件路径，确保目录存在
                temp_path = f"{image_path}_temp_resized.png"
                temp_dir = os.path.dirname(temp_path)
                
                # 如果目录不存在，尝试创建（但通常应该存在，因为原文件在这里）
                if not os.path.exists(temp_dir):
                    try:
                        os.makedirs(temp_dir, exist_ok=True)
                    except Exception as e:
                        print(f"无法创建临时文件目录 {temp_dir}: {e}")
                        return image_path  # 返回原路径作为备选
                
                # 保存临时文件
                resized_img.save(temp_path, "PNG")
                
                # 验证临时文件是否成功创建
                if os.path.exists(temp_path):
                    return temp_path
                else:
                    print(f"临时文件创建失败: {temp_path}")
                    return image_path
                
        except Exception as e:
            print(f"调整图片大小失败 {image_path}: {e}")
            import traceback
            traceback.print_exc()
            return image_path
    
    def _format_lora_info(self, lora_info_str: str) -> str:
        """格式化LoRA信息"""
        if not lora_info_str:
            return ""
        
        try:
            from .data_manager import DataManager
            dm = DataManager()
            lora_info = dm._deserialize_lora_info(lora_info_str)
            
            if not lora_info:
                return ""
            
            if isinstance(lora_info, dict) and 'loras' in lora_info:
                lora_list = []
                for lora in lora_info['loras']:
                    if isinstance(lora, dict):
                        name = lora.get('name', '未知')
                        weight = lora.get('weight', 1.0)
                        lora_list.append(f"{name} ({weight})")
                return "; ".join(lora_list)
            elif isinstance(lora_info, dict):
                # 字典格式
                lora_list = [f"{name} ({weight})" for name, weight in lora_info.items()]
                return "; ".join(lora_list)
            else:
                return str(lora_info)
                
        except Exception as e:
            print(f"格式化LoRA信息失败: {e}")
            return "解析错误"
    
    def _apply_styles(self, include_images: bool):
        """应用样式"""
        # 设置列宽
        column_widths = []
        
        if include_images:
            column_widths.append(25)  # 图片列，增加宽度
            
        column_widths.extend([
            15,  # 文件名
            15,  # 自定义名称
            40,  # 正向提示词
            30,  # 负向提示词
            20,  # 模型
            15,  # 采样器
            10,  # 采样步数
            10,  # CFG缩放
            18,  # 种子，增加宽度
            25,  # LoRA信息
            20,  # 标签
            25,  # 备注
            18   # 创建时间
        ])
        
        # 应用列宽
        for col, width in enumerate(column_widths, 1):
            self.worksheet.column_dimensions[get_column_letter(col)].width = width
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用边框到所有使用的单元格
        max_row = self.worksheet.max_row
        max_col = self.worksheet.max_column
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                self.worksheet.cell(row=row, column=col).border = thin_border
        
        # 冻结首行
        self.worksheet.freeze_panes = "A2"
    
    def _cleanup_temp_files(self, temp_files: List[str]):
        """清理临时文件"""
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    print(f"已清理临时文件: {temp_file}")
            except Exception as e:
                print(f"清理临时文件失败 {temp_file}: {e}")